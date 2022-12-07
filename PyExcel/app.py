import openpyxl

wb = openpyxl.load_workbook("transactions.xlsx")
print(wb.sheetnames)

sheet = wb["Sheet1"]


# access cells:
cell = sheet["A1"]
print(cell.value)
print(cell.row, cell.column)
print(cell.coordinate)


cell = sheet.cell(row=1, column=1)
# print(cell)

print(sheet.max_row)
print(sheet.max_column)

for row in range(1, sheet.max_row+1):
    for col in range(1, sheet.max_column+1):
        cell = sheet.cell(row, col)
        print(cell.value)


columns = sheet["a"]
print(columns)


cols = sheet["a:c"]
print(cols)

sheet.append([1, 2, 3])
wb.save("transactions2.xlsx")
